from typing import (Dict, List, Any, Type, cast, )
# from warnings import deprecated  # python >= 3.13

from PySide6.QtCore import (
    QAbstractTableModel, 
    )

from openpyxl import (Workbook, load_workbook, )
from openpyxl.styles import PatternFill, Font, fills, colors
from openpyxl.utils.datetime import from_excel, WINDOWS_EPOCH

from .cQModels import (SQLAlchemyTableModel, )
from ..database import Repository

ExcelWorkbook_fileext = ".XLSX"

class cExcelFile(Workbook):
    """Subclass of openpyxl's Workbook to add custom functionality for Excel file handling."""
    
    num_rows = None
    nRows = None
    nRowsSkipped = None
    
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Additional initialization can be added here if needed
    

    def load_from_listofdict(self,
        qlist:List[Dict[str, Any]], 
        wsName:str|None = None,
        freezecols:int = 0, 
        ) -> bool:
        """
        Load data from a SQLAlchemy model into the workbook.
        
        qset: a QAbstractTableModel or list of dictionaries
        flName: the name of the file to be built (WITHOUT extension!).  It's stored on the server.  If it's to be dl'd, the caller does that
        freezecols = 0: the number of columns to freeze to the left
        The top row contains the field names, is always frozen, is bold and is shaded grey

        Returns True if the workbook was successfully populated from the qset, False otherwise.
        """
        
        # create empty worksheet
        wb = self
        baseName = wsName or "Sheet"
        use_parens = not wsName
        candidate = baseName
        n = 0
        while candidate in wb.sheetnames:
            n += 1
            candidate = f"{baseName}({n})" if use_parens else f"{baseName}{n}"
        ws = wb.create_sheet(title=candidate)

        # header row is names of columns
        if qlist:
            fields = list(qlist[0])
            ws.append(fields)

            # append each row
            for row in qlist:
                ws.append(list(row.values())) # type: ignore

            # make header row bold, shade it grey, freeze it
            # ws.show_gridlines = True  #Nope - this is a R/O attribute
            for cell in ws[1]: # type: ignore
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type=fills.FILL_SOLID,
                                start_color=colors.Color("00808080"),
                                end_color=colors.Color("00808080")
                                )
            #TODO: convert row1 and cols:freezecols to an address (A=0, B=1, C=2 etc) for line below
            ws.freeze_panes ='A2'
            #TODO: if freezecols passed, freeze them, too
        #endif qlist

        return True
    # load_from_listofdict
    
    def load_from_SQLAlchModel(self,
        qset:SQLAlchemyTableModel, 
        wsName:str|None = None,
        freezecols:int = 0, 
        ) -> bool:
        """
        qset: a QAbstractTableModel or list of dictionaries
        flName: the name of the file to be built (WITHOUT extension!).  It's stored on the server.  If it's to be dl'd, the caller does that
        freezecols = 0: the number of columns to freeze to the left
        The top row contains the field names, is always frozen, is bold and is shaded grey

        Returns True if the workbook was successfully populated from the qset, False otherwise.
        """

        # far easier to process a list of dictionaries, so...
        if isinstance(qset,QAbstractTableModel):
            # make this a util
            qlist = qset.getDataAsList()
        elif isinstance(qset,list):
            qlist = qset
        else:
            return False
        if qlist:
            if not isinstance(qlist[0],dict):
                # review this later ...
                try:
                    qlist = [n.__dict__ for n in qlist]
                except:
                    qlist = []

        return self.load_from_listofdict(qlist, wsName, freezecols)
    # load_from_SQLAlchModel

    @classmethod
    def load_from_file(cls, filename: str, *args, **kwargs) -> "cExcelFile | None":
        """Load an existing Excel file and return it as cExcelFile.

        Args:
            filename (str): The name of the file to load, including extension.
            For other arguments, see openpyxl.reader.excel.load_workbook() documentation.

        Returns:
            Loaded cExcelFile instance, or None on failure.
        """
        try:
            wb = load_workbook(filename, *args, **kwargs)
            wb.__class__ = cls
            return cast("cExcelFile", wb)
        except Exception as e:
            print(f"Error loading file {filename}: {e}")
            return None
    # load_from_file

# move this into a method of cExcelFile
# class UpldSprdsheet():
    """Base class for handling spreadsheet uploads with field validation.
    
    This class provides functionality for processing uploaded spreadsheets,
    validating field types, and cleaning data according to defined rules.
    
    Attributes:
        TargetModel: The target ORM model class for the spreadsheet data.
        SprdsheetDateEpoch: Date epoch used for spreadsheet date conversion.
        SprdsheetFlds (dict): Dictionary mapping spreadsheet field names to field descriptors.
    
    This is now implemented in save_to_SQLAlchemyModel method of cExcelFile, which uses the SprdsheetFlds dict 
    TODO: create save_to_listofdict method 
    
    """
    # shoudl I just go on and make this a dataclass?
    class SprdsheetFldDescriptor:
        """Descriptor for spreadsheet fields, defining validation rules and cleaning procedures."""
        """Create a field descriptor for spreadsheet field validation.
        
        Args:
            ModelFldName (str): The name of the field in the TargetModel.
            AllowedTypes: tuple (or singleton) specifying allowed types
                for the field value. If not provided, defaults to str.
            CleanProc: Optional cleaning procedure to apply to the field value.
        
        Returns:
            dict: Field descriptor dictionary with ModelFldName and AllowedTypes.
        """
        ModelFldName:str
        CalculatedFld:bool
        AllowedTypes:tuple
        CleanProc:Any

        def __init__(self, ModelFldName:str, CalculatedFld=None, AllowedTypes=None, CleanProc=None):
            self.ModelFldName:str = ModelFldName
            self.CalculatedFld:bool = CalculatedFld or False
            self.AllowedTypes:tuple = AllowedTypes if AllowedTypes is not None else (str, )
            self.CleanProc = CleanProc
      # key will be the SprdsheetName, value is a SprdsheetFldDescriptor
    SprdsheetFlds:Dict[str,SprdsheetFldDescriptor] = {}  # key will be the SprdsheetName, value is a SprdsheetFldDescriptor
      # key will be the SprdsheetName, value is a SprdsheetFldDescriptor

    def cleanupfld(self, spshtFldNm, dbFldNm, val, rowdict):
        """Clean and validate a field value according to its allowed types.
        
        Args:
            fld: Field name to clean.
            val: Value to clean and validate.
            row: The entire row of data, which may be used in the cleaning procedure.
        
        Returns:
            tuple: (usefld, cleanval) where usefld is True if the field should be used,
                and cleanval is the cleaned value.
        """
        usefld = False
        cleanval = val
        
        if spshtFldNm not in self.SprdsheetFlds:
            # just feed the value back
            usefld = True
            cleanval = val
            return usefld, cleanval
        # endif fld not in self.SprdsheetFlds
        
        cleanDesc = self.SprdsheetFlds[spshtFldNm]
        if callable(cleanDesc.CleanProc):
            result = cleanDesc.CleanProc(spshtFldNm, dbFldNm, val, rowdict)
            if isinstance(result, tuple) and len(result) == 2:
                usefld, cleanval = result
                if not usefld:
                    return usefld, cleanval
            else:
                usefld = (result is not None)
                cleanval = result if usefld else None
        # if callable(cleanDesc.CleanProc)
        
        allowed_types = cleanDesc.AllowedTypes
        if not isinstance(allowed_types, tuple) or allowed_types == ():
            allowed_types = str
        usefld = isinstance(cleanval, allowed_types)
        
        return usefld, cleanval
    # cleanupfld

    def save_to_SQLAlchemyModel(self, 
        ssnmaker, 
        TargetModel:Type[Any],
        WksheetName, # default to None?
        SprdsheetFlds:Dict[str,SprdsheetFldDescriptor]|None = None, 
        required_columns=None,      # these are model field names, not spreadsheet column names - they are mapped to spreadsheet column names via SprdsheetFlds
        progress_interval=100,
        progress_callback = None,
        validation_callback = None,
        ) -> bool:
      # key will be the SprdsheetName, value is a SprdsheetFldDescriptor
        """Process the imported spreadsheet data and save it to the database.
        
        Args:
            dbsession: Database session for saving the processed data.
            SprsheetName: Name of the spreadsheet being processed.
            data: The data extracted from the spreadsheet to be processed.
            
        Note:
            This method is not yet implemented and should be overridden in subclasses
            to provide specific processing logic for the imported data.
        """
        wb = self
        
        if SprdsheetFlds is not None:
            self.SprdsheetFlds = SprdsheetFlds
        
        ws = wb[WksheetName] if WksheetName in wb.sheetnames else wb.active
        if ws is None:
            print(f"Worksheet {WksheetName} not found in workbook.")
            return False
                
        SprshtHdrRow = ws[1]
        required_columns = required_columns if required_columns is not None else []
        dbFld_to_sprshtCol:dict[str, Any] = {req_col: None for req_col in required_columns}
        for col in SprshtHdrRow:
            if col.value in self.SprdsheetFlds:
                colval = str(col.value)
                modlFldNm = self.SprdsheetFlds[colval].ModelFldName
                colIndx = (col.column or 0) - 1
                # has this field already been mapped?
                if modlFldNm in dbFld_to_sprshtCol and dbFld_to_sprshtCol[modlFldNm] is not None:
                    # yes, that's not good - duplicate column in spreadsheet.  Warn and skip this column
                    print(f"Warning: Duplicate column '{colval}' in spreadsheet. Column {dbFld_to_sprshtCol[modlFldNm]+1} Already mapped to model field '{modlFldNm}'. Skipping this column ({colIndx+1}).")
                    continue
                dbFld_to_sprshtCol[modlFldNm] = colIndx
        # end for col in SprshtHdrRow
        
        if any(val is None for val in dbFld_to_sprshtCol.values()):
            missing_cols = [col for col, val in dbFld_to_sprshtCol.items() if val is None]
            print(f"Error: Missing required columns in spreadsheet: {', '.join(missing_cols)}")
            return False
        # end if any missing required columns
        
        # Process each row in the worksheet
        self.num_rows = ws.max_row
        self.nRows = 0
        self.nRowsSkipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.nRows += 1
            if self.nRows % progress_interval == 0:
                print(f"Processing row {self.nRows} of {self.num_rows}...")
                if callable(progress_callback):
                    progress_callback(self.nRows, self.num_rows)
            # if nRows % progress_interval == 0
            
            # convert row to dict for easier processing - keys are model field names, values are cell values
            rowdict =  {modlFldNm: row[colIndx] for modlFldNm, colIndx in dbFld_to_sprshtCol.items() if colIndx is not None and colIndx < len(row)}
            
            # does this row have all required columns?
            if any([rowdict.get(req_col, None) in (None, '') for req_col in required_columns]):
                self.nRowsSkipped += 1
                continue
            
            record_data = {}
            # process the row
            for dbFldNm, val in rowdict.items():
                spshtFldNm = next((fld for fld, desc in self.SprdsheetFlds.items() if desc.ModelFldName == dbFldNm), None)
                if spshtFldNm:
                    usefld, cleanval = self.cleanupfld(spshtFldNm, dbFldNm, val, rowdict)
                    if usefld:
                        record_data[dbFldNm] = cleanval
            # end for dbFld, colIndx in dbFld_to_SsprshtCol.items()
            # process calculated fields, if any
            for spshtFldNm, desc in self.SprdsheetFlds.items():
                if not desc.CalculatedFld:
                    continue
                dbFldNm = desc.ModelFldName
                usefld, cleanval = self.cleanupfld(spshtFldNm, dbFldNm, None, rowdict)
                if usefld:
                    record_data[dbFldNm] = cleanval
            # end for fld_name, desc in self.SprdsheetFlds.items()
            
            validRecord = True            
            if callable(validation_callback):
                validRecord = validation_callback(record_data)

            # create an instance of TargetModel with record_data and add it to the session
            if validRecord:
                new_record = TargetModel(**record_data)
                Repository(ssnmaker, TargetModel).add(new_record)
            else:
                self.nRowsSkipped += 1
                print(f"Row {self.nRows} skipped due to validation failure.\n    {record_data}")
            # endif validRecord
        # for each row
        
        return True
    # save_to_SQLAlchemyModel

    def save_to_listofdict(self, 
        WksheetName, # default to None?
        SprdsheetFlds:Dict[str,SprdsheetFldDescriptor]|None = None, 
        required_columns=None,      # these are model field names, not spreadsheet column names - they are mapped to spreadsheet column names via SprdsheetFlds
        progress_interval=100,
        progress_callback = None,
        validation_callback = None,
        ) -> List[Dict[str, Any]]:
      # key will be the SprdsheetName, value is a SprdsheetFldDescriptor
        """Process the imported spreadsheet data and save it to the database.
        
        Args:
            dbsession: Database session for saving the processed data.
            SprsheetName: Name of the spreadsheet being processed.
            data: The data extracted from the spreadsheet to be processed.
            
        Note:
            This method is not yet implemented and should be overridden in subclasses
            to provide specific processing logic for the imported data.
        """
        wb = self
        
        if SprdsheetFlds is not None:
            self.SprdsheetFlds = SprdsheetFlds
        
        ws = wb[WksheetName] if WksheetName in wb.sheetnames else wb.active
        if ws is None:
            print(f"Worksheet {WksheetName} not found in workbook.")
            return [{'error': f"Worksheet {WksheetName} not found in workbook."}]
                
        SprshtHdrRow = ws[1]
        required_columns = required_columns if required_columns is not None else []
        dbFld_to_sprshtCol:dict[str, Any] = {req_col: None for req_col in required_columns}
        for col in SprshtHdrRow:
            if col.value in self.SprdsheetFlds:
                colval = str(col.value)
                modlFldNm = self.SprdsheetFlds[colval].ModelFldName
                colIndx = (col.column or 0) - 1
                # has this field already been mapped?
                if modlFldNm in dbFld_to_sprshtCol and dbFld_to_sprshtCol[modlFldNm] is not None:
                    # yes, that's not good - duplicate column in spreadsheet.  Warn and skip this column
                    print(f"Warning: Duplicate column '{colval}' in spreadsheet. Column {dbFld_to_sprshtCol[modlFldNm]+1} Already mapped to model field '{modlFldNm}'. Skipping this column ({colIndx+1}).")
                    continue
                dbFld_to_sprshtCol[modlFldNm] = colIndx
        # end for col in SprshtHdrRow
        
        if any(val is None for val in dbFld_to_sprshtCol.values()):
            missing_cols = [col for col, val in dbFld_to_sprshtCol.items() if val is None]
            print(f"Error: Missing required columns in spreadsheet: {', '.join(missing_cols)}")
            return [{'error': f"Missing required columns in spreadsheet: {', '.join(missing_cols)}"}]
        # end if any missing required columns
        
        data_list = []
        
        # Process each row in the worksheet
        self.num_rows = ws.max_row
        self.nRows = 0
        self.nRowsSkipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.nRows += 1
            if self.nRows % progress_interval == 0:
                print(f"Processing row {self.nRows} of {self.num_rows}...")
                if callable(progress_callback):
                    progress_callback(self.nRows, self.num_rows)
            # if nRows % progress_interval == 0
            
            # convert row to dict for easier processing - keys are model field names, values are cell values
            rowdict =  {modlFldNm: row[colIndx] for modlFldNm, colIndx in dbFld_to_sprshtCol.items() if colIndx is not None and colIndx < len(row)}
            
            # does this row have all required columns?
            if any([rowdict.get(req_col, None) in (None, '') for req_col in required_columns]):
                self.nRowsSkipped += 1
                continue
            
            record_data = {}
            # process the row
            for dbFldNm, val in rowdict.items():
                spshtFldNm = next((fld for fld, desc in self.SprdsheetFlds.items() if desc.ModelFldName == dbFldNm), None)
                if spshtFldNm:
                    usefld, cleanval = self.cleanupfld(spshtFldNm, dbFldNm, val, rowdict)
                    if usefld:
                        record_data[dbFldNm] = cleanval
            # end for dbFld, colIndx in dbFld_to_SsprshtCol.items()
            # process calculated fields, if any
            for spshtFldNm, desc in self.SprdsheetFlds.items():
                if not desc.CalculatedFld:
                    continue
                dbFldNm = desc.ModelFldName
                usefld, cleanval = self.cleanupfld(spshtFldNm, dbFldNm, None, rowdict)
                if usefld:
                    record_data[dbFldNm] = cleanval
            # end for fld_name, desc in self.SprdsheetFlds.items()
            
            validRecord = True            
            if callable(validation_callback):
                validRecord = validation_callback(record_data)

            # create an instance of TargetModel with record_data and add it to the session
            if validRecord:
                data_list.append(record_data)
            else:
                self.nRowsSkipped += 1
                print(f"Row {self.nRows} skipped due to validation failure.\n    {record_data}")
                data_list.append({
                    f'error-{self.nRows}': f"Row {self.nRows} skipped due to validation failure.", 
                    'incoming row': rowdict,
                    'processed row': record_data
                    })
            # endif validRecord
        # for each row
        
        return data_list
    # save_to_SQLAlchemyModel

# cExcelFile
    def end_of_class(self):
        """Placeholder method to indicate the end of the cExcelFile class definition."""
        pass
